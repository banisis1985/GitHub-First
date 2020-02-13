import openpyxl as xl
import glob, os, time
from datetime import datetime
from openpyxl.utils import column_index_from_string
from openpyxl import load_workbook
from pathlib import Path
while True:
    time.sleep(10)  #nurodome nesibaigianti cikla
    try:
        all_files = glob.glob('d:/vbaniukaitis/Desktop/codingCsv/dataIN/*.xlsx') #failo vieta, galima ir kitaip :)
        latest_xlsx = max(all_files, key=os.path.getctime) #paima paskutini faila is direktorijos
        wb = xl.load_workbook(latest_xlsx) #excel komandos
        ws = wb.active
        mylist= []
        datestring = datetime.strftime(datetime.now(), ' %Y_%m_%d %H_%M_%S') #pridedame data prie issaugoto failo
        for column in range(1, ws.max_column + 1): #skenuojam pavadinima, pagal norima pavadinima
            if ws.cell(1, column).value == 'ExpirationDate':
                stulpelis = column
        for row in range(2, ws.max_row +1):  #skenuojam visas eilutes
            mylist.append(ws.cell(row,stulpelis).value)
            for index in mylist: #isvengiam klaidu skenuojant
                index = str(index)
                if index != datetime:
                    continue
            if index.isalpha():
                continue
            date = datetime.strptime(index, '%Y-%m-%d %H:%M:%S') #skenuojam data pagal uzduota datetime
            newdate = date.replace(hour=23, minute=59) #pakeiciam valandas i norima reiksme
            cell = ws.cell(row, stulpelis)
            cell.value = (newdate) #irasome naujas reiksmes i nauja faila
        wb.save('d:/vbaniukaitis/Desktop/codingCsv/naujas' + datestring + '.xlsx') #naujo failo direktorija
    except:
        print('klaida') #esant klaidai
    continue
print('buvo nuskaitytas failas ' +latest_xlsx)
